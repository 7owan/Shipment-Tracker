import os
import sys
import requests
import time
from xml.etree import ElementTree
from openpyxl import load_workbook
from dotenv import load_dotenv
from concurrent.futures import ThreadPoolExecutor, as_completed
from threading import Lock
from datetime import datetime

load_dotenv()

aduie_lock = Lock()
last_aduie_time = 0

def print_progress_bar(iteration, total, length=40):
    percent = f"{100 * (iteration / float(total)):.1f}"
    filled_length = int(length * iteration // total)
    bar = '█' * filled_length + '-' * (length - filled_length)
    sys.stdout.write(f'\rProgress: |{bar}| {percent}% ({iteration}/{total})')
    sys.stdout.flush()

def clean_cell(cell):
    return str(cell.value or "").strip()

def get_manitoulin_auth_token():
    url = "https://www.mtdirect.ca/api/users/auth"
    payload = {
        "username": MANITOULIN_USERNAME,
        "token": MANITOULIN_LONG_TOKEN,
        "company": "MANITOULIN"
    }
    response = requests.post(url, json=payload)
    response.raise_for_status()
    return response.json().get("token")

def get_fedex_auth_token(client_id, client_secret):
    url = "https://apis.fedex.com/oauth/token"
    headers = {"Content-Type": "application/x-www-form-urlencoded"}
    data = {
        "grant_type": "client_credentials",
        "client_id": client_id,
        "client_secret": client_secret
    }
    response = requests.post(url, headers=headers, data=data)
    response.raise_for_status()
    return response.json()["access_token"]

def get_ups_auth_token(client_id, client_secret):
    url = "https://onlinetools.ups.com/security/v1/oauth/token"
    headers = {"Content-Type": "application/x-www-form-urlencoded"}
    data = {"grant_type": "client_credentials"}
    response = requests.post(url, headers=headers, auth=(client_id, client_secret), data=data)
    response.raise_for_status()
    return response.json()["access_token"]

def format_tstcf_date(date_str):
    """Convert YYYYMMDD to YYYY-MM-DD"""
    return datetime.strptime(date_str, "%Y%m%d").strftime("%Y-%m-%d")

def get_delivery_date_tstcf(requestor, authorization, tracking_number):
    """
    Calls TST-CF Express API and returns the delivery date if delivered.
    """
    url = "https://www.tst-cfexpress.com/xml/tracing"
    
    xml_payload = f"""<?xml version="1.0" encoding="ISO-8859-1"?>
<tracingrequest>
  <requestor>{requestor}</requestor>
  <authorization>{authorization}</authorization>
  <language>en</language>
  <tracetype>P</tracetype>
  <traceitems>
    <item>{tracking_number}</item>
  </traceitems>
</tracingrequest>
"""

    headers = {"Content-Type": "text/xml; charset=ISO-8859-1"}

    try:
        response = requests.post(url, data=xml_payload.encode("ISO-8859-1"), headers=headers, timeout=15)
        response.raise_for_status()
        
        root = ElementTree.fromstring(response.text)
        
        for traceitem in root.findall(".//traceitem"):
            valid = traceitem.findtext("valid")
            if valid != "Y":
                return None
            delivery = traceitem.find("delivery")
            if delivery is not None:
                date_el = delivery.find("date")
                if date_el is not None and date_el.text:
                    return format_tstcf_date(date_el.text)
    except Exception as e:
        return f"TSTCF error: {e}"

    return None
def format_tstcf_date(date_str):
    """Convert YYYYMMDD to YYYY-MM-DD"""
    return datetime.strptime(date_str, "%Y%m%d").strftime("%Y-%m-%d")

def get_delivery_date_tstcf(requestor, authorization, tracking_number):
    """
    Calls TST-CF Express API and returns the delivery date if delivered.
    """
    url = "https://www.tst-cfexpress.com/xml/tracing"
    
    xml_payload = f"""<?xml version="1.0" encoding="ISO-8859-1"?>
<tracingrequest>
  <requestor>{requestor}</requestor>
  <authorization>{authorization}</authorization>
  <language>en</language>
  <tracetype>P</tracetype>
  <traceitems>
    <item>{tracking_number}</item>
  </traceitems>
</tracingrequest>
"""

    headers = {"Content-Type": "text/xml; charset=ISO-8859-1"}

    try:
        response = requests.post(url, data=xml_payload.encode("ISO-8859-1"), headers=headers, timeout=15)
        response.raise_for_status()
        
        root = ElementTree.fromstring(response.text)
        
        for traceitem in root.findall(".//traceitem"):
            valid = traceitem.findtext("valid")
            if valid != "Y":
                return None
            delivery = traceitem.find("delivery")
            if delivery is not None:
                date_el = delivery.find("date")
                if date_el is not None and date_el.text:
                    return format_tstcf_date(date_el.text)
    except Exception as e:
        return f"TSTCF error: {e}"

    return None

def format_loomis_date(date_str):
    return datetime.strptime(date_str, "%Y%m%d").strftime("%Y-%m-%d")

def get_delivery_date_loomis(tracking_number):
    url = "https://webservice.loomis-express.com/LShip/services/USSAddonsService.USSAddonsServiceHttpsSoap12Endpoint/"

    soap_body = f"""<?xml version="1.0" encoding="utf-8"?>
    <soapenv:Envelope xmlns:soapenv="http://schemas.xmlsoap.org/soap/envelope/"
                      xmlns:ws="http://ws.addons.uss.transforce.ca"
                      xmlns:xsd="http://ws.addons.uss.transforce.ca/xsd">
       <soapenv:Header/>
       <soapenv:Body>
          <ws:trackByBarcode>
             <ws:request>
                <xsd:barcode>{tracking_number}</xsd:barcode>
                <xsd:track_shipment>false</xsd:track_shipment>
             </ws:request>
          </ws:trackByBarcode>
       </soapenv:Body>
    </soapenv:Envelope>
    """

    headers = {
        "Content-Type": "text/xml;charset=UTF-8",
        "SOAPAction": ""
    }

    try:
        response = requests.post(url, data=soap_body, headers=headers)
        response.raise_for_status()
        root = ElementTree.fromstring(response.text)
        ns = {
            'ax23': 'http://dto.uss.transforce.ca/xsd'
        }
        for event in root.findall('.//ax23:events', ns):
            code = event.find('ax23:code', ns)
            if code is not None and code.text in ("DEL", "NSR", "SGR", "DES"):
                date_time = event.find('ax23:local_date_time', ns)
                if date_time is not None:
                    return format_loomis_date(date_time.text[:8])
    except Exception:
        return None
    return None

def get_delivery_date_fedex(access_token, tracking_number):
    url = "https://apis.fedex.com/track/v1/trackingnumbers"
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {access_token}"
    }
    payload = {
        "trackingInfo": [{"trackingNumberInfo": {"trackingNumber": tracking_number}}],
        "includeDetailedScans": False
    }
    try:
        response = requests.post(url, headers=headers, json=payload)
        response.raise_for_status()
        result = response.json()
        shipment = result["output"]["completeTrackResults"][0]["trackResults"][0]
        for date in shipment.get("dateAndTimes", []):
            if date["type"] in ["ACTUAL_DELIVERY", "ESTIMATED_DELIVERY"]:
                return date["dateTime"][:10]
    except Exception:
        return None
    return None

def get_delivery_date_aduiepyle(user_email, tracking_number):
    global last_aduie_time
    url = "https://api.aduiepyle.com/2/shipment/status"
    params = {
        'user': user_email,
        'type': 0,
        'value': tracking_number
    }

    with aduie_lock:
        elapsed = time.time() - last_aduie_time
        wait_time = max(2.5 - elapsed, 0)
        if wait_time > 0:
            time.sleep(wait_time)
        last_aduie_time = time.time()

        try:
            response = requests.get(url, params=params)
            if response.status_code == 200:
                root = ElementTree.fromstring(response.text)
                for status in root.findall(".//statusDetail"):
                    if status.find("description").text == "DELIVERED":
                        return status.find("start").text[:10]
        except Exception:
            return None
    return None

def format_ups_date(date_str):
    return datetime.strptime(date_str, "%Y%m%d").strftime("%Y-%m-%d")

def get_delivery_date_ups(access_token, tracking_number):
    url = f"https://onlinetools.ups.com/api/track/v1/details/{tracking_number}"
    headers = {
        "Content-Type": "application/json",
        "Authorization": f"Bearer {access_token}",
        "transId": "track123",
        "transactionSrc": "testing"
    }
    try:
        response = requests.get(url, headers=headers)
        response.raise_for_status()
        data = response.json()
        activities = data.get("trackResponse", {}).get("shipment", [])[0].get("package", [])[0].get("activity", [])
        for activity in activities:
            if activity.get("status", {}).get("type") == "D":
                return format_ups_date(activity["date"])
    except Exception:
        return None
    return None

def get_delivery_date_manitoulin(token, probill_number):
    url = f"https://www.mtdirect.ca/api/probill/search/{probill_number}"
    headers = {
        "Authorization": f"Token {token}",
        "Content-Type": "application/json"
    }
    try:
        response = requests.get(url, headers=headers)
        if response.status_code == 200:
            data = response.json()
            delivered_on = data.get("details", {}).get("delivered_on", "")
            date_only = delivered_on.split(" at ")[0] if " at " in delivered_on else delivered_on
            return date_only if date_only else None
        else:
            print(f"[-] Manitoulin tracking failed for {probill_number}: {response.status_code}")
    except Exception as e:
        print(f"[-] Manitoulin tracking error for {probill_number}: {e}")
    return None

def track_package(row_idx, row, col_indices, fedex_token, ad_email, ups_token, manitoulin_token):
    carrier = clean_cell(row[col_indices["Carrier"]]).upper()
    tracking = clean_cell(row[col_indices["Tracking"]])
    delivered_value = clean_cell(row[col_indices["Delivered"]])

    if delivered_value or not carrier or not tracking:
        return row_idx, None

    try:
        if carrier in ["FEP", "FEE", "FEU", "FEA", "FED", "FEC"]:
            date = get_delivery_date_fedex(fedex_token, tracking)
        elif carrier == "DUE":
            date = get_delivery_date_aduiepyle(ad_email, tracking)
        elif carrier in ["UPS", "BLU", "UPO", "RED"]:
            date = get_delivery_date_ups(ups_token, tracking)
        elif carrier == "MAN":
            if not manitoulin_token:
                date = None
            else:
                date = get_delivery_date_manitoulin(manitoulin_token, tracking)
        elif carrier in ["LOO", "LAI"]:
            date = get_delivery_date_loomis(tracking)
        elif carrier == "TST":
            date = get_delivery_date_tstcf(TSTCF_REQUESTOR, TSTCF_AUTHORIZATION, tracking)
        else:
            date = None
    except Exception as e:
        date = f"Tracking error: {e}"

    return row_idx, date

def process_tracking_sheet(filename, fedex_token, ad_email, ups_token, manitoulin_token, sheet_name="Sheet1"):
    wb = load_workbook(filename)
    ws = wb.active

    header_row = [cell.value.strip() if cell.value else "" for cell in ws[1]]
    col_indices = {"Carrier": None, "Tracking": None, "Delivered": None}

    for idx, header in enumerate(header_row):
        header_lower = header.lower()
        if header_lower == "carrier":
            col_indices["Carrier"] = idx
        elif header_lower in ["pro #", "pro number", "tracking"]:
            col_indices["Tracking"] = idx
        elif header_lower == "delivered date":
            col_indices["Delivered"] = idx

    if None in col_indices.values():
        print("Missing required column(s):", col_indices)
        return

    rows = list(ws.iter_rows(min_row=2))
    total_rows = len(rows)
    updated_rows = 0

    with ThreadPoolExecutor(max_workers=10) as executor:
        futures = []
        for i, row in enumerate(rows):
            futures.append(executor.submit(track_package, i, row, col_indices, fedex_token, ad_email, ups_token, manitoulin_token))

        progress_lock = Lock()
        completed = 0

        for f in as_completed(futures):
            idx, result = f.result()
            if result:
                rows[idx][col_indices["Delivered"]].value = result
                updated_rows += 1
            with progress_lock:
                completed += 1
                print_progress_bar(completed, total_rows)

    print()
    wb.save(filename)
    print(f"Done. Updated '{filename}' with {updated_rows} new delivery dates.")

if __name__ == "__main__":
    start_time = time.time()

    FEDEX_CLIENT_ID = os.getenv("FEDEX_CLIENT_ID")
    FEDEX_CLIENT_SECRET = os.getenv("FEDEX_CLIENT_SECRET")
    ADUIEPYLE_EMAIL = os.getenv("ADUIEPYLE_EMAIL")
    UPS_CLIENT_ID = os.getenv("UPS_CLIENT_ID")
    UPS_CLIENT_SECRET = os.getenv("UPS_CLIENT_SECRET")
    MANITOULIN_USERNAME = os.getenv("MANITOULIN_USERNAME")
    MANITOULIN_LONG_TOKEN = os.getenv("MANITOULIN_LONG_TOKEN")
    TSTCF_REQUESTOR = os.getenv("TSTCF_REQUESTOR")
    TSTCF_AUTHORIZATION = os.getenv("TSTCF_AUTHORIZATION")

    try:
        FEDEX_TOKEN = get_fedex_auth_token(FEDEX_CLIENT_ID, FEDEX_CLIENT_SECRET)
    except Exception as e:
        print("FedEx auth error:", e)
        FEDEX_TOKEN = None

    try:
        UPS_TOKEN = get_ups_auth_token(UPS_CLIENT_ID, UPS_CLIENT_SECRET)
    except Exception as e:
        print("UPS auth error:", e)
        UPS_TOKEN = None

    MANITOULIN_TOKEN = get_manitoulin_auth_token()

    for file_name in os.listdir():
        if file_name.endswith(".xlsx"):
            try:
                print(f"Processing: {file_name}")
                process_tracking_sheet(file_name, fedex_token=FEDEX_TOKEN, ad_email=ADUIEPYLE_EMAIL, ups_token=UPS_TOKEN, manitoulin_token=MANITOULIN_TOKEN)
            except Exception as e:
                print(f"Error processing {file_name}: {e}")

    elapsed_time = time.time() - start_time
    minutes, seconds = divmod(elapsed_time, 60)
    print(f"Total time taken: {int(minutes)}m {int(seconds)}s")
